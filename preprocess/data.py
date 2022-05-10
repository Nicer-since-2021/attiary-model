import openpyxl
import random
from openpyxl import Workbook, load_workbook
from kobert_transformers import get_tokenizer


def wellness_question_data():
    root_path = "../data"
    wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
    wellness_q_output = root_path + "/wellness_dialog_question.txt"

    f = open(wellness_q_output, 'w')
    wb = load_workbook(filename=wellness_file)
    ws = wb[wb.sheetnames[0]]

    first_line = 0
    for row in ws.iter_rows():
        if first_line == 0:
            first_line = 1
            continue
        f.write(row[0].value + "    " + row[1].value + "\n")

    f.close()


def wellness_answer_data():
    root_path = "../data"
    wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
    wellness_a_output = root_path + "/wellness_dialog_answer.txt"

    f = open(wellness_a_output, 'w')
    wb = load_workbook(filename=wellness_file)
    ws = wb[wb.sheetnames[0]]

    first_line = 0

    for row in ws.iter_rows():
        if first_line == 0:
            first_line = 1
            continue

        if row[2].value == None:
            continue
        else:
            f.write(row[0].value + "    " + row[2].value + "\n")
    f.close()


def wellness_category_data():
    root_path = "../data"
    wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
    wellness_c_output = root_path + "/wellness_dialog_category.txt"

    f = open(wellness_c_output, 'w')
    wb = load_workbook(filename=wellness_file)
    ws = wb[wb.sheetnames[0]]

    first_line = 0
    category_count = 0
    cate_dict = []
    for row in ws.iter_rows():
        if first_line == 0:
            first_line = 1
            continue
        a = row[0].value
        if a not in cate_dict:
            cate_dict.append(a)
            f.write(row[0].value + "    " + str(category_count) + "\n")
            category_count += 1

    f.close()


def wellness_text_classification_data():
    root_path = "../data"
    wellness_category_file = root_path + "/wellness_dialog_category.txt"
    wellness_question_file = root_path + "/wellness_dialog_question.txt"
    wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"

    cate_file = open(wellness_category_file, 'r')
    ques_file = open(wellness_question_file, 'r')
    text_classfi_file = open(wellness_text_classification_file, 'w')

    category_lines = cate_file.readlines()
    cate_dict = {}
    for line_num, line_data in enumerate(category_lines):
        data = line_data.split('    ')
        cate_dict[data[0]] = data[1][:-1]

    ques_lines = ques_file.readlines()
    ques_dict = {}
    for line_num, line_data in enumerate(ques_lines):
        data = line_data.split('    ')
        text_classfi_file.write(data[1][:-1] + "    " + cate_dict[data[0]] + "\n")

    cate_file.close()
    ques_file.close()
    text_classfi_file.close()


def seperate_wellness_data():
    root_path = "../data"
    file_path = root_path + "/wellness_dialog_for_text_classification.txt"
    train_file_path = root_path + "/wellness_dialog_for_text_classification_train.txt"
    test_file_path = root_path + "/wellness_dialog_for_text_classification_test.txt"

    sperated_file = open(file_path, 'r')
    train_file = open(train_file_path, 'w')
    test_file = open(test_file_path, 'w')

    sperated_file_lines = sperated_file.readlines()
    ques_dict = {}
    for line_num, line_data in enumerate(sperated_file_lines):
        rand_num = random.randint(0, 10)
        if rand_num < 10:
            train_file.write(line_data)
        else:
            test_file.write(line_data)

    sperated_file.close()
    train_file.close()
    test_file.close()


if __name__ == '__main__':
    wellness_question_data()
    wellness_answer_data()
    wellness_category_data()
    wellness_text_classification_data()
    seperate_wellness_data()
